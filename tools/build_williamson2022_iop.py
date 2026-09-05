# -*- coding: utf-8 -*-
"""Extract the a and b spectra of Williamson & Hollins (2022).

The published Table 7 has 10 nm resolution; the accompanying spreadsheet has
1 nm. This reads the spreadsheet and checks the result against the printed
table and against the paper's own scattering equations.
"""

import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from _common import DATA_DIR, require, report

import openpyxl, csv, numpy as np
wb=openpyxl.load_workbook(require("20290782/20221121-Dstl_MIOP_analysis_v3.xlsx"),data_only=True)
ws=wb['a,b JIB-5C']
T=["IB","II","III","1C","3C","5C"]
data={}   # (qty, type) -> {wl: value}
wls=[]
for r in range(6, ws.max_row+1):
    w=ws.cell(r,2).value
    if not isinstance(w,(int,float)): continue
    w=int(w); wls.append(w)
    for i,t in enumerate(T):
        a=ws.cell(r,3+i).value; b=ws.cell(r,9+i).value
        if a is not None: data.setdefault(("a",t),{})[w]=float(a)
        if b is not None: data.setdefault(("b",t),{})[w]=float(b)
print(f"read {len(wls)} wavelengths, {min(wls)}-{max(wls)} nm, spacing {sorted(set(np.diff(sorted(wls))))}")

# Check 1: does it agree with the printed Table 7 (10 nm spacing)?
PUB={"a":{300:[0.149,0.219,0.349,0.566,0.572,1.91],400:[0.0448,0.0745,0.13,0.186,0.236,0.469],
          500:[0.0398,0.0529,0.0835,0.104,0.144,0.199],600:[0.227,0.231,0.241,0.248,0.268,0.285],
          700:[0.573,0.575,0.581,0.585,0.598,0.614],800:[2.29,2.29,2.29,2.29,2.3,2.29]},
     "b":{300:[0.178,0.258,0.374,0.531,0.791,1.45],400:[0.144,0.209,0.324,0.474,0.713,1.32],
          500:[0.128,0.185,0.297,0.44,0.664,1.23],600:[0.118,0.17,0.278,0.416,0.627,1.16],
          700:[0.111,0.159,0.264,0.396,0.599,1.11],800:[0.105,0.151,0.253,0.381,0.575,1.07]}}
print("\ncheck 1: against the printed Table 7")
bad=[]
for q in ("a","b"):
    for w,vals in PUB[q].items():
        for i,t in enumerate(T):
            v=data[(q,t)].get(w)
            if v is None or abs(v-vals[i])/vals[i]>0.005: bad.append((q,t,w,v,vals[i]))
print("  mismatches:", bad if bad else "none (72 points, all within 0.5%)")

# Check 2: is b reproduced by Eqs. (7)-(11) with a small-particle coefficient of 1.1513?
PAR={"IB":(0.010,0.37),"II":(0.022,0.52),"III":(0.00,0.90),"1C":(0.02,1.32),"3C":(0.00,2.07),"5C":(0.00,3.8)}
print("\ncheck 2: Eqs. (7)-(11) with Table 6 Bs, Bl and coefficient 1.1513")
print(f"{'type':>5} {'Bs':>7} {'Bl':>6} {'median %':>10} {'max %':>8}")
for t in T:
    Bs,Bl=PAR[t]; ws_=np.array(sorted(data[("b",t)]),float)
    obs=np.array([data[("b",t)][int(w)] for w in ws_])
    pred=0.00583*(400/ws_)**4.322 + Bs*1.1513*(400/ws_)**1.7 + Bl*0.341074*(400/ws_)**0.3
    e=100*(pred-obs)/obs
    print(f"{t:>4} {Bs:>7} {Bl:>6} {np.median(e):>+10.2f} {e[np.argmax(abs(e))]:>+9.2f}")

# Write the table.
n=0
with open(DATA_DIR / "williamson2022_iop.csv","w",newline="",encoding="utf-8") as f:
    wr=csv.writer(f)
    wr.writerow(["water_type","wavelength_nm","quantity","value_per_m","unit","status","note"])
    for t in T:
        for w in sorted(wls):
            for q in ("a","b"):
                v=data[(q,t)].get(w)
                if v is None: continue
                st,nt="ok",""
                if w<412 or w>715:
                    st="model_extrapolation"
                    nt="measured points exist only from 412 to 715 nm; this wavelength is model interpolation or extrapolation"
                wr.writerow([t,w,q,f"{v:.6g}","1/m",st,nt]); n+=1
report("williamson2022_iop.csv", n)
